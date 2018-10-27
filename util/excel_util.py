import os
import csv
import glob
from xlsxwriter.workbook import Workbook
import openpyxl as op
from util import filename_util


def generate_xlsx_from_csv(dir_path):
    for csvfile in glob.glob(os.path.join(dir_path, '*.csv')):
        workbook = Workbook(csvfile[:-4] + '.xlsx')
        worksheet = workbook.add_worksheet()
        with open(csvfile, 'rt', encoding='utf8') as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader):
                for c, col in enumerate(row):
                    worksheet.write(r, c, col)

        workbook.close()


# merge xlsx file from path_array
def generate_merged_xlsx_file(path_array, target_merged_path):
    merged_workbook = op.Workbook()

    for path in path_array:
        # load workbook
        workbook = op.load_workbook(filename=path)

        file_name = filename_util.extract_filename(path)

        # load all worksheets and copy the worksheets
        for worksheet in workbook.worksheets:
            new_sheet = merged_workbook.create_sheet(file_name + "-" + worksheet.title)

            # append ws1 to ws2
            for row in worksheet:
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value

    # create merged file with ws2
    merged_workbook.save(target_merged_path)
