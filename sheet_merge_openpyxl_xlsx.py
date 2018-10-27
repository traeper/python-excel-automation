#
# Example) merge one sheet
#
import openpyxl as op

path1 = 'xlsx/samsung-payment.xlsx'
path2 = 'xlsx/lg-payment.xlsx'
path3 = 'xlsx/merged.xlsx'

wb1 = op.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = op.load_workbook(filename=path2)
ws2 = wb2.create_sheet('new_sheet')

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path3)