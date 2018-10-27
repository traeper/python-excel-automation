import openpyxl as op

wb = op.load_workbook('xlsx/samsung-payment.xlsx')
sheet = wb['Sheet1']

sheets = wb.sheetnames

print(sheet['A1'].value)

sheet['A1'].value = 43000

# single cell
print(sheet.cell(row=1, column=1).value)

# multiple cell
print(sheet['A1':'A14'])

# first of multiple cell
print(sheet['A1':'A14'][0][0].value)
